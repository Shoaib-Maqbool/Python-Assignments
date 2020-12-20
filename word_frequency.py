# -*- coding: utf-8 -*-
"""
Created on Mon Aug 31 19:36:27 2020

@author: Shoib
"""
import nltk
from nltk.tokenize import word_tokenize,sent_tokenize
from nltk.probability import FreqDist
from openpyxl import load_workbook
wb = load_workbook("large.xlsx")  # Work Book
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
column = ws['A']  # Column
column_list = [column[x].value for x in range(len(column))]
from nltk.corpus import stopwords
nltk.download('stopwords')
stopwords_ = set(stopwords.words('english'))
stopwords_ = list(stopwords_)
nltk.download('wordnet')
from nltk.stem.wordnet import WordNetLemmatizer
lemmatizer = WordNetLemmatizer() 


while(None in column_list): 
    column_list.remove(None) 

tokenized_words = []
for i in column_list:
    i = word_tokenize(i)
    for j in i:
        j.lower()
        if j not in stopwords_ and len(j) > 2:
          
            lemmatizer.lemmatize(j)
            tokenized_words.append(j)

word_dist = FreqDist(tokenized_words)
print(word_dist.most_common(15))